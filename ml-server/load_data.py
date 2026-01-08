"""
order-ai.xlsx의 English 시트를 읽어서 SQLite DB에 저장
"""

import sqlite3
import openpyxl
import os

def load_english_sheet_to_db():
    """English 시트 → SQLite DB 변환"""
    
    # 경로 설정
    xlsx_path = os.path.join(os.path.dirname(__file__), "..", "order-ai.xlsx")
    db_path = os.path.join(os.path.dirname(__file__), "..", "data.sqlite3")
    
    if not os.path.exists(xlsx_path):
        print(f"❌ Excel 파일을 찾을 수 없습니다: {xlsx_path}")
        return
    
    print(f"📖 Excel 파일 읽기: {xlsx_path}")
    
    # Excel 읽기
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    if "English" not in wb.sheetnames:
        print("❌ 'English' 시트를 찾을 수 없습니다")
        print(f"   사용 가능한 시트: {wb.sheetnames}")
        return
    
    sheet = wb["English"]
    print(f"✅ 'English' 시트 발견 (행: {sheet.max_row})")
    
    # DB 연결
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 테이블 생성 (items가 없으면)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ml_items (
            item_no TEXT PRIMARY KEY,
            item_name TEXT NOT NULL,
            korean_name TEXT,
            english_name TEXT,
            vintage TEXT,
            country TEXT,
            producer TEXT,
            region TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    # 데이터 삽입
    inserted = 0
    skipped = 0
    
    for row_idx in range(2, sheet.max_row + 1):  # 2번 행부터 (헤더 제외)
        item_no = sheet.cell(row_idx, 2).value  # B열: 품목코드
        english_name = sheet.cell(row_idx, 8).value  # H열: 영문명
        korean_name = sheet.cell(row_idx, 9).value  # I열: 한글명
        vintage = sheet.cell(row_idx, 10).value  # J열: 빈티지
        country = sheet.cell(row_idx, 4).value  # D열: 국가
        producer = sheet.cell(row_idx, 5).value  # E열: 생산자
        region = sheet.cell(row_idx, 6).value  # F열: 지역
        
        # 필수 필드 체크
        if not item_no or (not english_name and not korean_name):
            skipped += 1
            continue
        
        # item_name 생성 (한글명 / 영문명 (빈티지))
        if korean_name and english_name:
            item_name = f"{korean_name} / {english_name}"
            if vintage:
                item_name += f" ({vintage})"
        elif korean_name:
            item_name = korean_name
        else:
            item_name = english_name
        
        try:
            cursor.execute("""
                INSERT OR REPLACE INTO ml_items 
                (item_no, item_name, korean_name, english_name, vintage, country, producer, region)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                str(item_no).strip(),
                item_name,
                korean_name,
                english_name,
                vintage,
                country,
                producer,
                region
            ))
            inserted += 1
        except Exception as e:
            print(f"⚠️ 행 {row_idx} 삽입 실패: {e}")
            skipped += 1
    
    conn.commit()
    conn.close()
    
    print(f"\n✅ 완료:")
    print(f"   - 삽입: {inserted}개")
    print(f"   - 스킵: {skipped}개")
    print(f"   - DB: {db_path}")

if __name__ == "__main__":
    load_english_sheet_to_db()
